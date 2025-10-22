import pdfplumber
import re
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

padrao_codigo = re.compile(r'\d{4}')
padrao_pagamento = re.compile(r'\d{2}/\d{2}/\d{4}')
padrao_valor = re.compile(r'\d{1,3}(?:\.\d{3})*,\d{2}')

def extracao_dados(arquivo, excel):
    aba = 'DADOS'
    dados_brutos = []
    data_referencia = ''
    codigo = ''
    data_pagamento = ''
    nome = ''
    ref = ''
    valor = ''

    with pdfplumber.open(arquivo) as pdf:
        for pagina in tqdm(pdf.pages, desc= 'Extraindo dados do arquivo', unit = 'Pagina', ncols = 100):
            texto = pagina.extract_text()
            linhas = texto.split('\n')
            for linha in linhas:
                if linha.startswith('Mês'):
                    match_mes = re.search(r'\d{1,2}', linha)
                    match_ano = re.search(r'\d{1}\.\d{3}', linha)

                    mes = int(match_mes.group())
                    ano = match_ano.group().replace('.','')
                    data_referencia = f'01/{mes:02d}/{ano}'

                match_codigo = padrao_codigo.match(linha)
                if match_codigo:
                    codigo = match_codigo.group()
                    aux_linha = linha[4:].strip()

                    match_valores = list(padrao_valor.finditer(linha))
                    if len(match_valores) >= 2:
                        v1 = match_valores[0]
                        v2 = match_valores[1]

                        ref = v1.group()
                        valor = v2.group()

                    match_pagamento = list(padrao_pagamento.finditer(aux_linha))
                    if len(match_pagamento) > 0:
                        p1 = match_pagamento[0]
                        data_pagamento = p1.group()

                        nome = aux_linha[:p1.start()].strip()

                dados_brutos.append([data_referencia, codigo, nome, data_pagamento, ref, valor])

    df = pd.DataFrame(dados_brutos, columns = ['DATA REFERENCIA', 'CODIGO', 'NOME', 'DATA PAGAMENTO', 'REF', 'VALOR'])
    df['DATA REFERENCIA'] = pd.to_datetime(df['DATA REFERENCIA'], format = '%d/%m/%Y')
    df['DATA PAGAMENTO'] = pd.to_datetime(df['DATA PAGAMENTO'], format='%d/%m/%Y')
    df['REF'] = (df['REF'].str.replace('.','',regex=False).str.replace(',','.',regex=False))
    df['REF'] = pd.to_numeric(df['REF'], errors='coerce')
    df['VALOR'] = (df['VALOR'].str.replace('.','',regex=False).str.replace(',','.',regex=False))
    df['VALOR'] = pd.to_numeric(df['VALOR'], errors='coerce')

    with pd.ExcelWriter(excel, mode = 'w') as writer:
        df.to_excel(writer, sheet_name = aba, index = False)
    
    formatar_data(excel, aba)

def formatar_data(excel, aba):
    wb = load_workbook(excel)
    ws = wb[aba]
    for col in ('A', 'D'):
        for cell in ws[col]:
            if cell.row == 1:
                continue
            cell.number_format = 'DD/MM/YYYY'
    wb.save(excel)

def main_holerite():
    caminho_arquivo = r'C:\Users\aluno\Desktop\SETEC\holerite.pdf'
    caminho_excel = caminho_arquivo.replace('.pdf', '.xlsx')

    extracao_dados(caminho_arquivo, caminho_excel)

main_holerite()